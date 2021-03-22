import pyperclip
import openpyxl
from docxtpl import DocxTemplate
import os
import comtypes.client
import time
import tkinter as tk

def clear_esc(event):
    entry.delete(0, 'end')
    entry.focus_set()

    
def create_word():
    wb = openpyxl.load_workbook('Данные\\Данные.xlsx')
    sheet = wb.active
    n = sheet.max_row
    print(n)
    save_location = entry.get()
    if not os.path.isdir(save_location):
        os.makedirs(save_location)

    if '\\' in save_location:
        save_location.replace('\\', '\\\\')   

    for i in range(n):
        if i < (n-1):
            company = sheet['A' + str(i + 2)].value
            adress = sheet['B' + str(i + 2)].value
            post = sheet['C' + str(i + 2)].value
            short_director = sheet['D' + str(i + 2)].value
            initials = sheet['E' + str(i +2 )].value
            doc = DocxTemplate('Данные\\Письмо_шаблон.docx')
            context = {'company': company, 'adress': adress, 'post': post,
                       'short_director': short_director, 'initials': initials}
            doc.render(context)
            file = company[company.index('"')+1:-1]
            save_location_word = save_location + f'\\Письмо [{file}].docx'
            doc.save(save_location_word)
            
    print ('Формирование Word-документов завершено')
    os.chdir(save_location)
    pyperclip.copy(os.getcwd())

def create_with_pdf():
    wb = openpyxl.load_workbook('Данные\\Данные.xlsx')
    sheet = wb.active
    n = sheet.max_row
    save_location = entry.get()
    if not os.path.isdir(save_location):
        os.makedirs(save_location)
    save_location_pdf = save_location + '\\PDF'
    if not os.path.isdir(save_location_pdf):
        os.mkdir(save_location_pdf)

    if '\\' in save_location:
        save_location.replace('\\', '\\\\')

##    ss = 0
    for i in range(n):
        if i < (n-1):
            company = sheet['A' + str(i + 2)].value
            adress = sheet['B' + str(i + 2)].value
            post = sheet['C' + str(i + 2)].value
            short_director = sheet['D' + str(i + 2)].value
            initials = sheet['E' + str(i +2 )].value
            doc = DocxTemplate('Данные\\Письмо_шаблон.docx')
            context = {'company': company, 'adress': adress, 'post': post,
                       'short_director': short_director, 'initials': initials}
            doc.render(context)
            file = company[company.index('"')+1:-1]
            save_location_word = save_location + f'\\Письмо [{file}].docx'
            doc.save(save_location_word)

            # создание pdf-файлов
            wdFormatPDF = 17
            
            in_file = save_location_word
            out_file = save_location_pdf + f'\\Письмо [{file}].pdf'
            
            word = comtypes.client.CreateObject('Word.Application')
            word.Visible = True
            time.sleep(1)

            doc=word.Documents.Open(in_file)
            doc.SaveAs(out_file, FileFormat=wdFormatPDF)
            doc.Close()
            word.Visible = False
##            ss += 1
##            if ss == 2:
##                break
        word.Quit()
        os.chdir(save_location)
        pyperclip.copy(os.getcwd())


win = tk.Tk()
win.geometry('370x120+900+400')
win.title('Формирование документов по шаблону')
win.resizable(0, 0)

enter = tk.Label(win, text='Введите адрес для сохранения документов (или имя папки):', anchor='w')
entry = tk.Entry(win)
btn_word = tk.Button (win, text='Создать Word-документы', command=create_word)
btn_pdf = tk.Button (win, text='Создать Word- и PDF-документы', command=create_with_pdf)
win.bind('<KeyPress-Escape>', clear_esc)


enter.place(x=8, y=5, width=365, height=20)
entry.place(x=5, y=30, width=360, height=30)
btn_word.place(x=5, y=70, width=160, height=40)
btn_pdf.place(x=170, y=70, width=195, height=40)

entry.focus_set()
win.mainloop()
